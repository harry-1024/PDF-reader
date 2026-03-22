import pdfplumber
import re
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


CAS_PATTERN = r'\b\d{2,7}-\d{2}-\d\b'


# ----------------------------
# Basic text helpers
# ----------------------------
def clean_chemical_name(name):
    if not name:
        return "UNKNOWN"
    name = re.sub(r'^[\d\.\-\s•*]+', '', name)
    name = name.split(';')[0].strip()
    return name


def normalize_text(text):
    return " ".join((text or "").split())


def make_product_key(product_name, product_id):
    return f"{(product_name or '').strip()} || {(product_id or '').strip()}"


def normalize_cas(cas):
    if cas is None:
        return ""
    return str(cas).strip()


# ----------------------------
# Averaging period lookup
# ----------------------------
def period_label_to_hours(label):
    label = (label or "").strip().lower()

    if "10-minute" in label or "10 minute" in label:
        return "0.167"
    if "1-hour" in label or "1 hour" in label:
        return "1"
    if "24-hour" in label or "24 hour" in label:
        return "24"
    if "monthly" in label or "month" in label:
        return "730"
    if "annual" in label or "year" in label:
        return "8760"

    return None


def extract_periods_from_concentration_cell(cell_value):
    text = str(cell_value or "").strip()

    matches = re.findall(r'\(([^)]*)\)', text, flags=re.IGNORECASE)
    periods = []

    for m in matches:
        converted = period_label_to_hours(m)
        if converted:
            periods.append(converted)

    if not periods:
        periods = ["24"]

    periods = sorted(set(periods), key=lambda x: float(x))
    return periods


def build_averaging_period_lookup(csv_path):
    if not os.path.exists(csv_path):
        print(f"⚠️ Averaging period CSV not found: {csv_path}")
        return {}

    df = pd.read_csv(csv_path)

    cas_col = None
    conc_col = None

    for c in df.columns:
        cl = str(c).strip().lower()
        if cl == "cas rn" or "cas" in cl:
            cas_col = c
        if "concentration" in cl:
            conc_col = c

    if cas_col is None or conc_col is None:
        print("⚠️ Could not identify CAS / concentration columns in ACB CSV.")
        return {}

    lookup = {}

    for _, row in df.iterrows():
        cas = normalize_cas(row.get(cas_col, ""))
        if not cas:
            continue

        periods = extract_periods_from_concentration_cell(row.get(conc_col, ""))

        if cas not in lookup:
            lookup[cas] = set()

        for p in periods:
            lookup[cas].add(p)

    final_lookup = {}
    for cas, period_set in lookup.items():
        sorted_periods = sorted(period_set, key=lambda x: float(x))
        final_lookup[cas] = ", ".join(sorted_periods)

    print(f"✅ Loaded averaging periods for {len(final_lookup)} CAS numbers from CSV.")
    return final_lookup


def get_averaging_period_for_cas(cas_number, averaging_lookup):
    cas_number = normalize_cas(cas_number)
    return averaging_lookup.get(cas_number, "24")


# ----------------------------
# Concentration parsing
# ----------------------------
def parse_conservative_percent(conc_text):
    if not conc_text:
        return None

    text = normalize_text(conc_text)

    m = re.search(
        r'(?:[≥>]\s*)?(\d+(?:\.\d+)?)\s*-\s*(?:[≤<]\s*)?(\d+(?:\.\d+)?)',
        text
    )
    if m:
        return float(m.group(2))

    m = re.search(r'[≤<≥>]?\s*(\d+(?:\.\d+)?)', text)
    if m:
        return float(m.group(1))

    return None


def extract_ingredient_from_line(line):
    if not line:
        return None

    line = normalize_text(line)
    cas_match = re.search(CAS_PATTERN, line)
    if not cas_match:
        return None

    cas = cas_match.group(0)
    left = line[:cas_match.start()].strip()

    conc_match = re.search(
        r'((?:[≥>]\s*)?\d+(?:\.\d+)?\s*-\s*(?:[≤<]\s*)?\d+(?:\.\d+)?|[≤<≥>]?\s*\d+(?:\.\d+)?)\s*$',
        left
    )
    if not conc_match:
        return None

    conc_text = conc_match.group(1).strip()
    name = left[:conc_match.start()].strip()
    percent = parse_conservative_percent(conc_text)

    if not name or percent is None:
        return None

    if not (0 <= percent <= 100):
        return None

    return {
        "CAS #": cas,
        "Contaminant": clean_chemical_name(name),
        "Percentage": percent
    }


# ----------------------------
# Section extraction
# ----------------------------
def extract_section_text(full_text, start_patterns, end_patterns):
    start_idx = None
    end_idx = None

    for pat in start_patterns:
        m = re.search(pat, full_text, flags=re.IGNORECASE)
        if m:
            start_idx = m.start()
            break

    if start_idx is None:
        return ""

    for pat in end_patterns:
        m = re.search(pat, full_text[start_idx:], flags=re.IGNORECASE)
        if m:
            end_idx = start_idx + m.start()
            break

    if end_idx is None:
        return full_text[start_idx:]
    return full_text[start_idx:end_idx]


def extract_section_1_text(full_text):
    return extract_section_text(
        full_text,
        start_patterns=[
            r'Section\s*1\.\s*Identification',
            r'SECTION\s*1\.\s*IDENTIFICATION',
            r'Section\s*1\b'
        ],
        end_patterns=[
            r'Section\s*2\.',
            r'SECTION\s*2\.',
            r'Section\s*2\b'
        ]
    )


def extract_section_3_text(full_text):
    return extract_section_text(
        full_text,
        start_patterns=[
            r'Section\s*3\.\s*Composition\s*/\s*information on ingredients',
            r'Section\s*3\.\s*Composition/information on ingredients',
            r'SECTION\s*3\.\s*COMPOSITION',
            r'Section\s*3\b'
        ],
        end_patterns=[
            r'Section\s*4\.\s*First aid measures',
            r'SECTION\s*4\.\s*FIRST AID',
            r'Section\s*4\b'
        ]
    )


def extract_section_9_text(full_text):
    return extract_section_text(
        full_text,
        start_patterns=[
            r'Section\s*9\.\s*Physical and chemical properties',
            r'SECTION\s*9\.\s*PHYSICAL AND CHEMICAL PROPERTIES',
            r'Section\s*9\b'
        ],
        end_patterns=[
            r'Section\s*10\.',
            r'SECTION\s*10\.',
            r'Section\s*10\b'
        ]
    )


# ----------------------------
# Product metadata extraction
# ----------------------------
def extract_product_name(section1_text, fallback_name):
    patterns = [
        r'Product name\s*[:\-]\s*(.+)',
        r'Product identifier\s*[:\-]\s*(.+)',
        r'Product Name\s*[:\-]\s*(.+)',
        r'Product Identifier\s*[:\-]\s*(.+)',
    ]

    for pat in patterns:
        m = re.search(pat, section1_text, flags=re.IGNORECASE)
        if m:
            value = m.group(1).strip().split('\n')[0].strip()
            if value:
                return value

    return fallback_name


def extract_product_id(section1_text):
    patterns = [
        r'Product code\s*[:\-]\s*(.+)',
        r'Product identification number\s*[:\-]\s*(.+)',
        r'Product Code\s*[:\-]\s*(.+)',
        r'Product Identification Number\s*[:\-]\s*(.+)',
        r'Product No\.\s*[:\-]\s*(.+)'
    ]

    for pat in patterns:
        m = re.search(pat, section1_text, flags=re.IGNORECASE)
        if m:
            value = m.group(1).strip().split('\n')[0].strip()
            if value:
                return value

    return ""


def extract_specific_gravity(section9_text, full_text):
    search_texts = [section9_text, full_text]

    patterns = [
        r'Specific gravity\s*[:\-]?\s*([0-9]+(?:\.[0-9]+)?)',
        r'Relative density\s*[:\-]?\s*([0-9]+(?:\.[0-9]+)?)',
        r'Specific Gravity\s*[:\-]?\s*([0-9]+(?:\.[0-9]+)?)',
        r'Relative Density\s*[:\-]?\s*([0-9]+(?:\.[0-9]+)?)'
    ]

    for text in search_texts:
        if not text:
            continue
        for pat in patterns:
            m = re.search(pat, text, flags=re.IGNORECASE)
            if m:
                return m.group(1).strip()

    return ""


# ----------------------------
# Flags
# ----------------------------
def detect_solids_flag(full_text):
    text = full_text.lower()
    solid_keywords = [
        "solid", "powder", "granules", "granular",
        "pellet", "crystal", "crystalline"
    ]
    return "s" if any(k in text for k in solid_keywords) else ""


# ----------------------------
# Parse one PDF
# ----------------------------
def parse_sds_file(pdf_path, averaging_lookup):
    fallback_name = os.path.splitext(os.path.basename(pdf_path))[0]

    with pdfplumber.open(pdf_path) as pdf:
        full_text = "\n".join([(p.extract_text() or "") for p in pdf.pages])

    section1_text = extract_section_1_text(full_text)
    section3_text = extract_section_3_text(full_text)
    section9_text = extract_section_9_text(full_text)

    product_name = extract_product_name(section1_text, fallback_name)
    product_id = extract_product_id(section1_text)
    specific_gravity = extract_specific_gravity(section9_text, full_text)
    solid_flag = detect_solids_flag(full_text)

    rows = []
    lines = [ln.strip() for ln in section3_text.split("\n") if ln.strip()]

    for line in lines:
        result = extract_ingredient_from_line(line)
        if result:
            avg_period = get_averaging_period_for_cas(result["CAS #"], averaging_lookup)

            rows.append({
                "Contaminant": result["Contaminant"],
                "CAS #": result["CAS #"],
                "Averaging Period(s)": avg_period if avg_period else "24",
                "Solid [1]": solid_flag,
                "Product Name": product_name,
                "Product Identification Number": product_id,
                "specific gravity": specific_gravity,
                "Percentage": result["Percentage"]
            })

    meta = {
        "Product Name": product_name,
        "Product Identification Number": product_id,
        "specific gravity": specific_gravity
    }

    return rows, meta


# ----------------------------
# Existing matrix reader
# ----------------------------
def load_existing_matrix_data(existing_file):
    """
    Backward compatible:
    - Old format:
      Col1 Contaminant | Col2 CAS # | Col3 Solid [1] | product starts at col5
    - New format:
      Col1 Contaminant | Col2 CAS # | Col3 Averaging Period(s) | Col4 Solid [1] | product starts at col6
    """
    if not os.path.exists(existing_file):
        return [], {}

    wb = load_workbook(existing_file, data_only=True)
    ws = wb.active

    row_product_name = 1
    row_product_id = 2
    row_spec_gravity = 3
    row_header = 4
    data_start_row = 5

    header_col3 = ws.cell(row=row_header, column=3).value
    has_avg_column = str(header_col3).strip().lower() == "averaging period(s)" if header_col3 else False

    if has_avg_column:
        product_start_col = 6
    else:
        product_start_col = 5

    product_meta_map = {}
    product_cols = []

    col = product_start_col
    while True:
        product_name = ws.cell(row=row_product_name, column=col).value
        if product_name in [None, ""]:
            break

        product_id = ws.cell(row=row_product_id, column=col).value or ""
        spec_gravity = ws.cell(row=row_spec_gravity, column=col).value or ""

        product_name = str(product_name).strip()
        product_id = str(product_id).strip()
        spec_gravity = str(spec_gravity).strip()

        product_meta_map[product_name] = {
            "Product Name": product_name,
            "Product Identification Number": product_id,
            "specific gravity": spec_gravity
        }
        product_cols.append((col, product_name))
        col += 1

    all_rows = []
    row = data_start_row
    while True:
        contaminant = ws.cell(row=row, column=1).value
        cas = ws.cell(row=row, column=2).value

        if contaminant in [None, ""] and cas in [None, ""]:
            break

        contaminant = "" if contaminant is None else str(contaminant).strip()
        cas = "" if cas is None else str(cas).strip()

        if has_avg_column:
            avg_period = ws.cell(row=row, column=3).value
            solid = ws.cell(row=row, column=4).value
        else:
            avg_period = "24"
            solid = ws.cell(row=row, column=3).value

        avg_period = "24" if avg_period in [None, ""] else str(avg_period).strip()
        solid = "" if solid is None else str(solid).strip()

        for col_idx, product_name in product_cols:
            cell_val = ws.cell(row=row, column=col_idx).value
            if cell_val not in [None, ""]:
                val_str = str(cell_val).replace("%", "").strip()
                try:
                    pct = float(val_str)
                except ValueError:
                    continue

                meta = product_meta_map[product_name]
                all_rows.append({
                    "Contaminant": contaminant,
                    "CAS #": cas,
                    "Averaging Period(s)": avg_period if avg_period else "24",
                    "Solid [1]": solid,
                    "Product Name": product_name,
                    "Product Identification Number": meta["Product Identification Number"],
                    "specific gravity": meta["specific gravity"],
                    "Percentage": pct
                })

        row += 1

    return all_rows, product_meta_map


def get_existing_product_keys(product_meta_map):
    keys = set()
    for meta in product_meta_map.values():
        keys.add(make_product_key(
            meta.get("Product Name", ""),
            meta.get("Product Identification Number", "")
        ))
    return keys


# ----------------------------
# Excel writer
# ----------------------------
def write_formatted_matrix(all_rows, product_meta_map, output_name, backup_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "SDS Matrix"

    blue_fill = PatternFill("solid", fgColor="0B63B5")
    white_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    df = pd.DataFrame(all_rows)

    # Standardize one contaminant name per CAS
    name_map = (
        df.groupby("CAS #")["Contaminant"]
        .apply(lambda x: max(x, key=len))
        .to_dict()
    )
    df["Contaminant"] = df["CAS #"].map(name_map)

    # Standardize averaging period by CAS
    def combine_avg_periods(series):
        valid_vals = []
        allowed_periods = {"0.167", "1", "24", "730", "8760"}

        for v in series:
            if v is None:
                valid_vals.append("24")
                continue

            raw = str(v).strip()
            if raw == "":
                valid_vals.append("24")
                continue

            parts = [x.strip() for x in raw.split(",") if x.strip()]

            found_valid = False
            for part in parts:
                if part in allowed_periods:
                    valid_vals.append(part)
                    found_valid = True

            # if cell contained only junk like "s", default to 24
            if not found_valid:
                valid_vals.append("24")

        valid_vals = sorted(set(valid_vals), key=lambda x: float(x))
        return ", ".join(valid_vals)

    avg_map = df.groupby("CAS #")["Averaging Period(s)"].apply(combine_avg_periods).to_dict()
    df["Averaging Period(s)"] = df["CAS #"].map(avg_map)

    products = list(product_meta_map.keys())

    matrix = df.pivot_table(
        index=["Contaminant", "CAS #", "Averaging Period(s)", "Solid [1]"],
        columns="Product Name",
        values="Percentage",
        aggfunc="max"
    ).reset_index()

    existing_products = [p for p in products if p in matrix.columns]
    matrix = matrix[["Contaminant", "CAS #", "Averaging Period(s)", "Solid [1]"] + existing_products]
    matrix = matrix.sort_values(by="Contaminant").reset_index(drop=True)

    row_product_name = 1
    row_product_id = 2
    row_spec_gravity = 3
    row_header = 4
    data_start_row = 5

    # Left header block: cols 1 to 5
    for r in [row_product_name, row_product_id, row_spec_gravity, row_header]:
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = blue_fill
            ws.cell(row=r, column=c).font = white_font
            ws.cell(row=r, column=c).alignment = center
            ws.cell(row=r, column=c).border = border

    ws.cell(row=row_product_name, column=5, value="Product Name")
    ws.cell(row=row_product_id, column=5, value="Product Identification Number")
    ws.cell(row=row_spec_gravity, column=5, value="specific gravity")

    ws.cell(row=row_header, column=1, value="Contaminant")
    ws.cell(row=row_header, column=2, value="CAS #")
    ws.cell(row=row_header, column=3, value="Averaging Period(s)")
    ws.cell(row=row_header, column=4, value="Solid [1]")

    start_col = 6
    for i, product in enumerate(existing_products, start=start_col):
        meta = product_meta_map.get(product, {})

        ws.cell(row=row_product_name, column=i, value=meta.get("Product Name", product))
        ws.cell(row=row_product_id, column=i, value=meta.get("Product Identification Number", ""))
        ws.cell(row=row_spec_gravity, column=i, value=meta.get("specific gravity", ""))

        for r in [row_product_name, row_product_id, row_spec_gravity, row_header]:
            ws.cell(row=r, column=i).fill = blue_fill
            ws.cell(row=r, column=i).font = white_font
            ws.cell(row=r, column=i).alignment = center
            ws.cell(row=r, column=i).border = border

    for row_idx, (_, row) in enumerate(matrix.iterrows(), start=data_start_row):
        ws.cell(row=row_idx, column=1, value=row["Contaminant"])
        ws.cell(row=row_idx, column=2, value=row["CAS #"])
        ws.cell(row=row_idx, column=3, value=row["Averaging Period(s)"] if row["Averaging Period(s)"] else "24")
        ws.cell(row=row_idx, column=4, value=row["Solid [1]"])

        for c in range(1, 5):
            ws.cell(row=row_idx, column=c).border = border
            ws.cell(row=row_idx, column=c).alignment = center

        for col_idx, product in enumerate(existing_products, start=start_col):
            val = row.get(product, "")
            if pd.notna(val) and val != "":
                if float(val).is_integer():
                    display_val = f"{int(val)}%"
                else:
                    display_val = f"{val}%"
            else:
                display_val = ""

            ws.cell(row=row_idx, column=col_idx, value=display_val)
            ws.cell(row=row_idx, column=col_idx).alignment = center
            ws.cell(row=row_idx, column=col_idx).border = border

    widths = {
        1: 38,
        2: 14,
        3: 18,
        4: 12,
        5: 28
    }
    for c, width in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = width

    for col_idx in range(start_col, start_col + len(existing_products)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22

    wb.save(output_name)
    wb.save(backup_name)


# ----------------------------
# Main
# ----------------------------
def process_all_sds_incremental(
    folder_path,
    output_name="Master_SDS_Matrix_Formatted.xlsx",
    acb_csv_name="mecp_acb_list.csv"
):
    backup_name = f"Master_SDS_Matrix_Formatted_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    if not os.path.exists(folder_path):
        print(f"Folder not found: {folder_path}")
        return

    csv_path = os.path.join(folder_path, acb_csv_name)
    averaging_lookup = build_averaging_period_lookup(csv_path)

    existing_rows, product_meta_map = load_existing_matrix_data(output_name)
    existing_product_keys = get_existing_product_keys(product_meta_map)

    print(f"📘 Existing products in matrix: {len(existing_product_keys)}")

    new_rows = []
    new_product_count = 0
    skipped_count = 0

    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]

    for filename in pdf_files:
        pdf_path = os.path.join(folder_path, filename)

        try:
            print(f"🔍 Checking: {filename}")
            rows, meta = parse_sds_file(pdf_path, averaging_lookup)

            product_key = make_product_key(
                meta.get("Product Name", ""),
                meta.get("Product Identification Number", "")
            )

            if product_key in existing_product_keys:
                print(f"⏭️ Skipping existing product: {meta.get('Product Name', filename)}")
                skipped_count += 1
                continue

            if not rows:
                print(f"⚠️ No ingredient rows found for {filename}")
                continue

            new_rows.extend(rows)
            product_meta_map[meta["Product Name"]] = meta
            existing_product_keys.add(product_key)
            new_product_count += 1
            print(f"➕ Added new product: {meta['Product Name']}")

        except Exception as e:
            print(f"❌ Error reading {filename}: {e}")

    combined_rows = existing_rows + new_rows

    if not combined_rows:
        print("❌ No data found.")
        return

    write_formatted_matrix(combined_rows, product_meta_map, output_name, backup_name)

    print("✅ Incremental update complete.")
    print(f"✅ New products added: {new_product_count}")
    print(f"✅ Existing products skipped: {skipped_count}")
    print(f"✅ Saved: {output_name}")
    print(f"✅ Backup: {backup_name}")


# ----------------------------
# Run
# ----------------------------
sds_folder = os.path.join(os.getcwd(), "sds_files")
process_all_sds_incremental(
    folder_path=sds_folder,
    output_name="Master_SDS_Matrix_Formatted.xlsx",
    acb_csv_name="mecp_acb_list.csv"
)
